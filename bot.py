"""
WhatsApp Web bulk sender for Windows.

Important:
- Use this only for contacts who gave permission to receive your messages.
- WhatsApp Web changes often. If a selector stops working, update the selector
  lists in the helper functions below.
- Keep Chrome visible. Headless mode is disabled by default because WhatsApp Web
  is much more reliable in a normal browser window.
"""

import os
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import pyautogui
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# =============================================================================
# EDITABLE CONFIG SECTION
# =============================================================================

MIN_DELAY = 5
MAX_DELAY = 120

DAILY_LIMIT = 100

ENABLE_ATTACHMENT = True
ATTACHMENT_PATH = "image.jpg"
SEND_TEXT_AS_ATTACHMENT_CAPTION = False

# Maximum 5 templates. The bot randomly chooses one for every contact.
TEMPLATES = [
    "Hi, I saw your business online.",
    "Hello, are you currently running ads?",
    "Hey, we help businesses grow online.",
    "Hi, do you currently have a website?",
    "Hello, I wanted to connect regarding digital marketing.",
]

TEMPLATE_NAMES = [
    "Default 1",
    "Default 2",
    "Default 3",
    "Default 4",
    "Default 5",
]

# Optional settings.
HEADLESS = False
RESUME_FROM_LOGS = False
CONTACTS_FILE = "contacts.xlsx"
TEMPLATES_FILE = "templates.txt"
LOGS_FILE = "logs.txt"
REPORT_FILE = "report.xlsx"
CHROME_DATA_DIR = "chrome-data"
CHAT_LOAD_TIMEOUT = 25
LOGIN_TIMEOUT = 300
CHAT_LOAD_RETRIES = 2

# =============================================================================
# END CONFIG SECTION
# =============================================================================


BASE_DIR = Path(__file__).resolve().parent
SUPPORTED_ATTACHMENT_TYPES = {".jpg", ".jpeg", ".png", ".pdf"}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# PyAutoGUI is included because it is useful for Windows automation fallback.
# The script mainly uses Selenium for stability, but this pause keeps any future
# PyAutoGUI action from firing too quickly.
pyautogui.PAUSE = 0.25
pyautogui.FAILSAFE = True


def print_info(message: str) -> None:
    print(f"[INFO] {message}", flush=True)


def print_success(message: str) -> None:
    print(f"[SUCCESS] {message}", flush=True)


def print_failed(message: str) -> None:
    print(f"[FAILED] {message}", flush=True)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")


def validate_config() -> None:
    """Fail early for config mistakes that are easy to fix."""
    if MIN_DELAY < 0 or MAX_DELAY < 0:
        raise ValueError("MIN_DELAY and MAX_DELAY must be positive numbers.")

    if MIN_DELAY > MAX_DELAY:
        raise ValueError("MIN_DELAY cannot be greater than MAX_DELAY.")

    if DAILY_LIMIT <= 0:
        raise ValueError("DAILY_LIMIT must be greater than zero.")

    if not TEMPLATES:
        raise ValueError("Add at least one message template.")

    if len(TEMPLATES) > 5:
        raise ValueError("Only 5 templates are allowed. Remove extra templates.")

    if ENABLE_ATTACHMENT:
        attachment = resolve_attachment_path(ATTACHMENT_PATH)
        extension = attachment.suffix.lower()
        if extension not in SUPPORTED_ATTACHMENT_TYPES:
            raise ValueError("Attachment must be jpg, png, or pdf.")
        if not attachment.exists():
            raise FileNotFoundError(f"Attachment not found: {attachment}")


def parse_template_block(block: str, fallback_name: str) -> tuple[str, str]:
    """Return a display name and sendable body from one templates.txt block."""
    lines = block.strip().splitlines()
    name = fallback_name

    while lines and not lines[0].strip():
        lines.pop(0)

    if lines:
        first_line = lines[0].strip()
        bracket_label = re.fullmatch(r"\[(.+?)\]", first_line)
        hash_label = re.fullmatch(r"#\s*(.+)", first_line)
        slash_label = re.fullmatch(r"//\s*(.+)", first_line)
        plain_label = re.fullmatch(r"(Template\s*\d+|Message\s*\d+)\s*[:\-]\s*(.+)", first_line, re.IGNORECASE)
        simple_template_label = re.fullmatch(r"(Template|Message)?\s*\d+\s*[\).\:-]?\s*(.*)", first_line, re.IGNORECASE)

        if bracket_label:
            name = bracket_label.group(1).strip()
            lines.pop(0)
        elif hash_label:
            name = hash_label.group(1).strip()
            lines.pop(0)
        elif slash_label:
            name = slash_label.group(1).strip()
            lines.pop(0)
        elif plain_label:
            name = f"{plain_label.group(1).strip()}: {plain_label.group(2).strip()}"
            lines.pop(0)
        elif simple_template_label:
            label_text = simple_template_label.group(0).strip()
            label_tail = simple_template_label.group(2).strip()
            name = label_tail or label_text
            lines.pop(0)

    body = "\n".join(lines).strip()
    return name, body


def load_templates_from_file() -> tuple[list[str], list[str]]:
    """Load multi-line templates from templates.txt while preserving line breaks."""
    templates_path = BASE_DIR / TEMPLATES_FILE
    if not templates_path.exists():
        return [], []

    content = templates_path.read_text(encoding="utf-8-sig").strip()
    if not content:
        return [], []

    raw_templates = [template.strip() for template in re.split(r"(?m)^\s*---TEMPLATE(?:\s+\d+)?---\s*$", content) if template.strip()]
    parsed_templates = [parse_template_block(template, f"Template {index}") for index, template in enumerate(raw_templates, start=1)]
    names = [name for name, body in parsed_templates if body]
    templates = [body for name, body in parsed_templates if body]

    if len(templates) > 5:
        raise ValueError(f"{TEMPLATES_FILE} contains {len(templates)} templates. Maximum allowed is 5.")

    return templates, names


def template_preview(template: str, max_length: int = 80) -> str:
    """Show a compact one-line preview without losing the real template text."""
    preview = " ".join(template.split())
    if len(preview) > max_length:
        return preview[: max_length - 3] + "..."
    return preview


def resolve_attachment_path(file_path: str) -> Path:
    """Support both full Windows paths and files placed beside bot.py."""
    cleaned = file_path.strip().strip('"').strip("'")
    path = Path(cleaned)
    if path.is_absolute():
        return path
    return BASE_DIR / path


def ask_float(prompt: str, default: float) -> float:
    """Ask for a decimal number. Empty input keeps the default."""
    while True:
        answer = input(f"{prompt} [{default}]: ").strip()
        if not answer:
            return float(default)
        try:
            value = float(answer)
            if value < 0:
                print_failed("Please enter a positive number.")
                continue
            return value
        except ValueError:
            print_failed("Please enter a valid number, example: 18.5")


def ask_int(prompt: str, default: int, minimum: int = 1, maximum: int | None = None) -> int:
    """Ask for a whole number. Empty input keeps the default."""
    while True:
        answer = input(f"{prompt} [{default}]: ").strip()
        if not answer:
            return int(default)
        try:
            value = int(answer)
        except ValueError:
            print_failed("Please enter a whole number.")
            continue

        if value < minimum:
            print_failed(f"Please enter at least {minimum}.")
            continue
        if maximum is not None and value > maximum:
            print_failed(f"Please enter no more than {maximum}.")
            continue
        return value


def ask_yes_no(prompt: str, default: bool) -> bool:
    """Ask a yes/no question. Empty input keeps the default."""
    default_text = "Y/n" if default else "y/N"
    while True:
        answer = input(f"{prompt} ({default_text}): ").strip().lower()
        if not answer:
            return default
        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False
        print_failed("Please type y or n.")


def collect_runtime_settings() -> None:
    """Guide the user through settings that commonly change every campaign."""
    global MIN_DELAY, MAX_DELAY, DAILY_LIMIT, ENABLE_ATTACHMENT, ATTACHMENT_PATH
    global TEMPLATES, TEMPLATE_NAMES, RESUME_FROM_LOGS

    print()
    print_info("Campaign setup")
    MIN_DELAY = ask_float("Minimum delay after each message in seconds", MIN_DELAY)
    MAX_DELAY = ask_float("Maximum delay after each message in seconds", MAX_DELAY)
    if MIN_DELAY > MAX_DELAY:
        print_info("Minimum delay was greater than maximum delay, so I swapped them.")
        MIN_DELAY, MAX_DELAY = MAX_DELAY, MIN_DELAY

    DAILY_LIMIT = ask_int("Daily sending limit", DAILY_LIMIT, minimum=1)
    RESUME_FROM_LOGS = ask_yes_no("Skip numbers already marked Sent in logs.txt", RESUME_FROM_LOGS)

    print()
    print_info("Message templates")
    file_templates, file_template_names = load_templates_from_file()
    if file_templates:
        print_info(f"Found {len(file_templates)} template(s) in {TEMPLATES_FILE}.")
        if ask_yes_no(f"Load templates from {TEMPLATES_FILE}", True):
            TEMPLATES = file_templates
            TEMPLATE_NAMES = file_template_names
    else:
        print_info(f"No {TEMPLATES_FILE} found. For multi-line templates, create that file beside bot.py.")

    for index, template in enumerate(TEMPLATES, start=1):
        name = TEMPLATE_NAMES[index - 1] if index <= len(TEMPLATE_NAMES) else f"Template {index}"
        print(f"  {index}. {name} - {template_preview(template)}")

    use_existing = ask_yes_no("Use these templates", True)
    if not use_existing:
        print_info(f"For messages with line breaks/emojis, edit {TEMPLATES_FILE} instead of typing here.")
        template_count = ask_int("How many templates do you want to add? Maximum is 5", 1, 1, 5)
        new_templates: list[str] = []
        for index in range(1, template_count + 1):
            while True:
                template = input(f"Enter template {index}: ").strip()
                if template:
                    new_templates.append(template)
                    break
                print_failed("Template cannot be empty.")
        TEMPLATES = new_templates
        TEMPLATE_NAMES = [f"Manual Template {index}" for index in range(1, len(TEMPLATES) + 1)]

    template_limit = ask_int(
        "How many templates do you want to use from the loaded list?",
        len(TEMPLATES),
        minimum=1,
        maximum=len(TEMPLATES),
    )
    TEMPLATES = TEMPLATES[:template_limit]
    TEMPLATE_NAMES = TEMPLATE_NAMES[:template_limit]
    print_info(f"Using {len(TEMPLATES)} template(s) for random selection.")

    print()
    ENABLE_ATTACHMENT = ask_yes_no("Send media/document with the message", ENABLE_ATTACHMENT)
    if ENABLE_ATTACHMENT:
        print_info("Supported files: jpg, png, pdf")
        while True:
            file_path = input(f"Attachment path [{ATTACHMENT_PATH}]: ").strip()
            if file_path:
                ATTACHMENT_PATH = file_path

            attachment = resolve_attachment_path(ATTACHMENT_PATH)
            if attachment.exists() and attachment.suffix.lower() in SUPPORTED_ATTACHMENT_TYPES:
                ATTACHMENT_PATH = str(attachment)
                print_info(f"Selected attachment: {attachment}")
                break

            print_failed("Attachment not found or unsupported. Use jpg, png, or pdf.")

    print()


def create_driver() -> webdriver.Chrome:
    """Open Chrome with a persistent profile so QR login is saved."""
    chrome_data = BASE_DIR / CHROME_DATA_DIR
    chrome_data.mkdir(exist_ok=True)

    options = Options()
    options.add_argument(f"--user-data-dir={chrome_data}")
    options.add_argument("--profile-directory=Default")
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument("--log-level=3")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    extension_dir = BASE_DIR / "extension"
    if (extension_dir / "manifest.json").exists():
        options.add_argument(f"--load-extension={extension_dir}")

    if HEADLESS:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1366,768")

    local_driver = BASE_DIR / "chromedriver.exe"
    if local_driver.exists():
        service = Service(str(local_driver), log_output=os.devnull)
    else:
        # Selenium Manager downloads/uses a compatible driver automatically when
        # Selenium 4.6+ is installed. Add chromedriver.exe only if this fails.
        service = Service(log_output=os.devnull)

    return webdriver.Chrome(service=service, options=options)


def read_contacts() -> list[str]:
    """Read and clean numbers from contacts.xlsx."""
    contacts_path = BASE_DIR / CONTACTS_FILE
    if not contacts_path.exists():
        raise FileNotFoundError(f"Missing contact file: {contacts_path}")

    df = pd.read_excel(contacts_path, dtype={"number": str})
    if "number" not in df.columns:
        raise ValueError('contacts.xlsx must contain a column named "number".')

    contacts: list[str] = []
    for raw_number in df["number"].dropna().tolist():
        number = re.sub(r"\s+", "", str(raw_number).strip())
        if number.endswith(".0"):
            number = number[:-2]
        contacts.append(number)

    return contacts


def is_basic_number_format_valid(number: str) -> bool:
    """Check local formatting before opening WhatsApp."""
    return bool(re.fullmatch(r"[1-9][0-9]{7,14}", number))


def load_sent_numbers_from_logs() -> set[str]:
    """Resume support: skip numbers already marked as sent in logs.txt."""
    sent_numbers: set[str] = set()
    logs_path = BASE_DIR / LOGS_FILE
    if not RESUME_FROM_LOGS or not logs_path.exists():
        return sent_numbers

    with logs_path.open("r", encoding="utf-8") as file:
        for line in file:
            parts = [part.strip() for part in line.split("|")]
            if len(parts) >= 3 and parts[2].lower() == "sent":
                sent_numbers.add(parts[1])

    return sent_numbers


def append_log(number: str, status: str, template_used: str, reason: str = "-") -> None:
    """Save every result in logs.txt as soon as it happens."""
    logs_path = BASE_DIR / LOGS_FILE
    line = f"{now_text()} | {number} | {status} | {template_used} | {reason}\n"
    with logs_path.open("a", encoding="utf-8") as file:
        file.write(line)


def clear_previous_queue_files() -> None:
    """Clear old run output files without touching contacts or templates."""
    logs_path = BASE_DIR / LOGS_FILE
    report_path = BASE_DIR / REPORT_FILE

    logs_path.write_text("", encoding="utf-8")
    if report_path.exists():
        report_path.unlink()

    print_info("Cleared old logs.txt and report.xlsx.")


def show_queue_details(contacts: list[str], sent_before: set[str]) -> set[str]:
    """Show queue summary and optionally clear old logs/report before sending."""
    while True:
        valid_count = sum(1 for number in contacts if is_basic_number_format_valid(number))
        invalid_format_count = len(contacts) - valid_count
        already_sent_count = sum(1 for number in contacts if number in sent_before)
        pending_count = max(0, valid_count - already_sent_count if RESUME_FROM_LOGS else valid_count)

        print()
        print_info("Queue Details")
        print(f"  Total contacts: {len(contacts)}")
        print(f"  Valid format: {valid_count}")
        print(f"  Invalid local format: {invalid_format_count}")
        print(f"  Already sent in logs: {already_sent_count}")
        print(f"  Pending for this run: {min(pending_count, DAILY_LIMIT)}")
        print(f"  Daily limit: {DAILY_LIMIT}")
        print()
        print("  [C] Clear old queue logs/report")
        print("  [ENTER] Continue")

        choice = input("Queue Details option: ").strip().lower()
        if choice == "c":
            clear_previous_queue_files()
            sent_before = set()
            continue
        if choice == "":
            return sent_before

        print_failed("Please press ENTER to continue or type C to clear.")

    return sent_before


def wait_for_login(driver: webdriver.Chrome) -> None:
    """Wait until WhatsApp Web shows the logged-in chat UI."""
    print_info("Opening WhatsApp Web...")
    driver.get("https://web.whatsapp.com/")

    qr_notice_shown = False
    end_time = time.time() + LOGIN_TIMEOUT

    while time.time() < end_time:
        if is_logged_in(driver):
            print_info("WhatsApp account detected successfully")
            return

        if not qr_notice_shown and is_qr_visible(driver):
            print_info("QR code detected. Scan it once with your phone.")
            qr_notice_shown = True

        time.sleep(1)

    raise TimeoutException("Login timeout. Scan QR and try again.")


def is_logged_in(driver: webdriver.Chrome) -> bool:
    """Return True when the WhatsApp chat list is visible."""
    login_selectors = [
        (By.ID, "pane-side"),
        (By.XPATH, "//div[@aria-label='Chat list']"),
        (By.XPATH, "//div[@role='grid' and @aria-label='Chat list']"),
    ]

    for by, selector in login_selectors:
        try:
            if driver.find_elements(by, selector):
                return True
        except WebDriverException:
            continue

    return False


def is_qr_visible(driver: webdriver.Chrome) -> bool:
    qr_selectors = [
        (By.XPATH, "//canvas[contains(@aria-label, 'Scan')]"),
        (By.XPATH, "//*[contains(text(), 'Scan this QR code')]"),
        (By.XPATH, "//*[contains(text(), 'Use WhatsApp on your computer')]"),
    ]

    for by, selector in qr_selectors:
        try:
            if driver.find_elements(by, selector):
                return True
        except WebDriverException:
            continue

    return False


def find_first_visible(driver: webdriver.Chrome, locators: list[tuple[str, str]], timeout: int = 10):
    """Try multiple selectors and return the first visible element."""
    wait = WebDriverWait(driver, timeout)
    last_error: Exception | None = None

    for by, selector in locators:
        try:
            return wait.until(EC.visibility_of_element_located((by, selector)))
        except Exception as error:
            last_error = error

    raise TimeoutException(str(last_error))


def click_first_available(driver: webdriver.Chrome, locators: list[tuple[str, str]], timeout: int = 10) -> None:
    """Try multiple selectors until one clickable element works."""
    wait = WebDriverWait(driver, timeout)
    last_error: Exception | None = None

    for by, selector in locators:
        try:
            element = wait.until(EC.element_to_be_clickable((by, selector)))
            element.click()
            return
        except (TimeoutException, ElementClickInterceptedException, WebDriverException) as error:
            last_error = error

    raise TimeoutException(str(last_error))


def is_invalid_number_page(driver: webdriver.Chrome) -> bool:
    """Detect WhatsApp's invalid number message."""
    invalid_phrases = [
        "phone number shared via url is invalid",
        "phone number is invalid",
        "invalid phone number",
        "not on whatsapp",
        "isn't on whatsapp",
        "is not on whatsapp",
        "number is not on whatsapp",
        "phone number is not on whatsapp",
        "phone number isn't on whatsapp",
        "not a valid whatsapp number",
        "not a valid phone number",
        "not registered on whatsapp",
        "couldn't find this phone number",
        "could not find this phone number",
        "no whatsapp account",
    ]

    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
        if any(phrase in body_text for phrase in invalid_phrases):
            return True
    except WebDriverException:
        pass

    invalid_selectors = [
        (
            By.XPATH,
            "//*[contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), "
            "'phone number shared via url is invalid')]",
        ),
        (
            By.XPATH,
            "//*[contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), "
            "'phone number is invalid')]",
        ),
        (
            By.XPATH,
            "//*[contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), "
            "'not on whatsapp')]",
        ),
    ]

    for by, selector in invalid_selectors:
        try:
            if driver.find_elements(by, selector):
                return True
        except WebDriverException:
            continue

    return False


def click_continue_to_chat_if_present(driver: webdriver.Chrome) -> None:
    """Click WhatsApp intermediate buttons if they appear before the chat opens."""
    continue_selectors = [
        (
            By.XPATH,
            "//*[self::button or @role='button'][contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue to chat')]",
        ),
        (
            By.XPATH,
            "//*[self::button or @role='button'][contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'use whatsapp web')]",
        ),
        (
            By.XPATH,
            "//*[self::button or @role='button'][contains(translate(normalize-space(.), "
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]",
        ),
    ]

    for by, selector in continue_selectors:
        try:
            elements = driver.find_elements(by, selector)
            for element in elements:
                if element.is_displayed() and element.is_enabled() and click_element_safely(driver, element):
                    time.sleep(1)
                    return
        except WebDriverException:
            continue


def wait_for_chat_box(driver: webdriver.Chrome, number: str):
    """Open a number and wait until the message input appears."""
    chat_url = f"https://web.whatsapp.com/send?phone={number}&type=phone_number&app_absent=0"

    box_selectors = [
        (By.XPATH, "//footer//div[@contenteditable='true' and @role='textbox']"),
        (By.XPATH, "//footer//*[@role='textbox']"),
        (By.XPATH, "//footer//div[@contenteditable='true' and @aria-label]"),
        (By.XPATH, "//footer//*[@contenteditable='true' and @data-lexical-editor='true']"),
        (By.XPATH, "//footer//*[@contenteditable='true']"),
        (By.XPATH, "//*[@aria-label='Type a message']"),
        (By.XPATH, "//*[@aria-placeholder='Type a message']"),
        (By.XPATH, "//*[contains(@aria-label, 'Type a message')]"),
        (By.XPATH, "//*[contains(@aria-placeholder, 'Type a message')]"),
        (By.XPATH, "//footer//p[contains(@class, 'selectable-text')]/ancestor::*[@contenteditable='true'][1]"),
    ]

    for attempt in range(1, CHAT_LOAD_RETRIES + 1):
        driver.get(chat_url)
        end_time = time.time() + CHAT_LOAD_TIMEOUT

        while time.time() < end_time:
            if is_invalid_number_page(driver):
                raise ValueError("Invalid Number")

            click_continue_to_chat_if_present(driver)

            try:
                return find_first_visible(driver, box_selectors, timeout=2)
            except TimeoutException:
                time.sleep(1)

        if is_invalid_number_page(driver):
            raise ValueError("Invalid Number")

        if attempt < CHAT_LOAD_RETRIES:
            print_info(f"Chat load retry {attempt}/{CHAT_LOAD_RETRIES} for {number}")

    raise TimeoutException("Chat did not load in time")


def set_windows_clipboard(text: str) -> None:
    """Copy text to the Windows clipboard using native Unicode clipboard APIs."""
    clipboard_text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\r\n")

    try:
        import ctypes

        cf_unicode_text = 13
        gmem_moveable = 0x0002
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32

        user32.OpenClipboard.argtypes = [ctypes.c_void_p]
        user32.OpenClipboard.restype = ctypes.c_bool
        user32.EmptyClipboard.argtypes = []
        user32.EmptyClipboard.restype = ctypes.c_bool
        user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
        user32.SetClipboardData.restype = ctypes.c_void_p
        user32.CloseClipboard.argtypes = []
        user32.CloseClipboard.restype = ctypes.c_bool
        kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
        kernel32.GlobalAlloc.restype = ctypes.c_void_p
        kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
        kernel32.GlobalLock.restype = ctypes.c_void_p
        kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
        kernel32.GlobalUnlock.restype = ctypes.c_bool

        clipboard_opened = False
        for _ in range(10):
            if user32.OpenClipboard(None):
                clipboard_opened = True
                break
            time.sleep(0.1)

        if not clipboard_opened:
            raise RuntimeError("OpenClipboard failed.")

        try:
            user32.EmptyClipboard()
            data = ctypes.create_unicode_buffer(clipboard_text)
            size = ctypes.sizeof(data)
            handle = kernel32.GlobalAlloc(gmem_moveable, size)
            if not handle:
                raise RuntimeError("GlobalAlloc failed.")

            locked_handle = kernel32.GlobalLock(handle)
            if not locked_handle:
                raise RuntimeError("GlobalLock failed.")

            ctypes.memmove(locked_handle, data, size)
            kernel32.GlobalUnlock(handle)

            if not user32.SetClipboardData(cf_unicode_text, handle):
                raise RuntimeError("SetClipboardData failed.")
        finally:
            user32.CloseClipboard()
    except Exception as error:
        raise RuntimeError(f"Could not set clipboard text: {short_error(error)}")


def normalize_message_text(text: str) -> str:
    """Normalize line endings only; keep message content intact."""
    return text.replace("\r\n", "\n").replace("\r", "\n")


def clear_message_box(driver: webdriver.Chrome, element) -> None:
    """Clear any WhatsApp draft before pasting a template."""
    try:
        driver.execute_script("arguments[0].focus();", element)
    except WebDriverException:
        pass

    element.click()
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)
    pyautogui.press("backspace")
    time.sleep(0.4)


def paste_text_into_box(driver: webdriver.Chrome, element, text: str) -> None:
    """Use real Windows copy/paste so the template stays exactly intact."""
    clean_text = normalize_message_text(text)
    expected_compact = re.sub(r"\s+", "", clean_text)

    set_windows_clipboard(clean_text)

    for attempt in range(1, 4):
        clear_message_box(driver, element)

        try:
            driver.execute_script("arguments[0].focus();", element)
        except WebDriverException:
            pass

        element.click()
        time.sleep(0.5)
        pyautogui.hotkey("ctrl", "v")
        time.sleep(1.8)

        pasted_text = normalize_message_text(get_editable_text(driver, element))
        pasted_compact = re.sub(r"\s+", "", pasted_text)

        if pasted_compact == expected_compact:
            print_info("Template pasted exactly from clipboard.")
            return

        if len(pasted_compact) >= max(5, int(len(expected_compact) * 0.95)):
            print_info("Template pasted from clipboard.")
            return

        print_info(f"Clipboard paste retry {attempt}/3")

    raise TimeoutException("Clipboard paste did not match template. Message was not sent.")


def get_editable_text(driver: webdriver.Chrome, element) -> str:
    """Read visible text from WhatsApp's contenteditable message boxes."""
    text = driver.execute_script(
        "return (arguments[0].innerText || arguments[0].textContent || '').trim();",
        element,
    )
    return "" if text is None else str(text).strip()


def click_element_safely(driver: webdriver.Chrome, element) -> bool:
    """Try normal click first, then JavaScript click for stubborn WhatsApp buttons."""
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
            element,
        )
        time.sleep(0.2)
    except WebDriverException:
        pass

    try:
        element.click()
        return True
    except WebDriverException:
        pass

    try:
        driver.execute_script("arguments[0].click();", element)
        return True
    except WebDriverException:
        return False


def send_button_selectors() -> list[tuple[str, str]]:
    """Selectors for WhatsApp's normal and attachment-preview Send buttons."""
    return [
        (By.XPATH, "//button[@aria-label='Send']"),
        (By.XPATH, "//*[@role='button' and @aria-label='Send']"),
        (By.XPATH, "//*[@data-testid='send']"),
        (By.XPATH, "//*[@data-icon='send']/ancestor::button[1]"),
        (By.XPATH, "//*[@data-icon='send']/ancestor::*[@role='button'][1]"),
        (By.XPATH, "//*[contains(@data-icon, 'send')]/ancestor::button[1]"),
        (By.XPATH, "//*[contains(@data-icon, 'send')]/ancestor::*[@role='button'][1]"),
        (By.XPATH, "//*[contains(@data-icon, 'wds-ic-send')]/ancestor::button[1]"),
        (By.XPATH, "//*[contains(@data-icon, 'wds-ic-send')]/ancestor::*[@role='button'][1]"),
        (By.XPATH, "//*[contains(@aria-label, 'Send') and (@role='button' or self::button)]"),
    ]


def click_send_button(driver: webdriver.Chrome, timeout: int = 20) -> None:
    """Click WhatsApp's Send button with several selector fallbacks."""

    end_time = time.time() + timeout
    last_error = "Send button not found"

    while time.time() < end_time:
        for by, selector in send_button_selectors():
            try:
                elements = driver.find_elements(by, selector)
            except WebDriverException as error:
                last_error = short_error(error)
                continue

            for element in elements:
                try:
                    if element.is_displayed() and element.is_enabled() and click_element_safely(driver, element):
                        return
                except WebDriverException as error:
                    last_error = short_error(error)

        time.sleep(0.5)

    raise TimeoutException(last_error)


def click_send_button_by_js_scan(driver: webdriver.Chrome, timeout: int = 20) -> None:
    """Find and click the most likely visible Send button using browser-side scan."""
    script = """
        const selectors = [
            'button[aria-label="Send"]',
            '[role="button"][aria-label="Send"]',
            '[data-testid="send"]',
            '[data-icon*="send"]',
            '[aria-label*="Send"]',
            '[title*="Send"]'
        ];

        const seen = new Set();
        const previewCandidates = [];
        const candidates = [];

        for (const selector of selectors) {
            for (const node of document.querySelectorAll(selector)) {
                const clickable = node.closest('button,[role="button"]') || node;
                if (seen.has(clickable)) continue;
                seen.add(clickable);

                const rect = clickable.getBoundingClientRect();
                const style = window.getComputedStyle(clickable);
                const visible = rect.width > 0 &&
                    rect.height > 0 &&
                    style.visibility !== 'hidden' &&
                    style.display !== 'none' &&
                    clickable.getAttribute('aria-disabled') !== 'true';

                if (visible) {
                    const inPreview =
                        Boolean(clickable.closest('[role="dialog"]')) ||
                        Boolean(clickable.closest('[aria-label*="Preview"]')) ||
                        Boolean(clickable.closest('[aria-label*="preview"]')) ||
                        Boolean(clickable.closest('[data-testid*="media"]')) ||
                        Boolean(clickable.closest('[data-testid*="drawer"]'));

                    const item = { element: clickable, top: rect.top, left: rect.left };
                    if (inPreview) {
                        previewCandidates.push(item);
                    }
                    candidates.push(item);
                }
            }
        }

        previewCandidates.sort((a, b) => (b.top - a.top) || (b.left - a.left));
        candidates.sort((a, b) => (b.top - a.top) || (b.left - a.left));

        const target = previewCandidates[0] || candidates[0];
        if (target) {
            target.element.click();
            return true;
        }

        return false;
    """

    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            if driver.execute_script(script):
                return
        except WebDriverException:
            pass
        time.sleep(0.5)

    raise TimeoutException("Send button JS scan failed.")


def click_attachment_send_by_screen_position(driver: webdriver.Chrome) -> None:
    """Last-resort click for WhatsApp's attachment preview send button."""
    rect = driver.get_window_rect()
    x = rect["x"] + rect["width"] - 95
    y = rect["y"] + rect["height"] - 110
    print_info(f"Trying preview Send button screen click at {x}, {y}")
    pyautogui.click(x, y)
    time.sleep(1.5)


def click_attachment_preview_send(driver: webdriver.Chrome, timeout: int = 25) -> None:
    """Click the Send button on the attachment preview using layered fallbacks."""
    errors: list[str] = []

    try:
        click_send_button_by_js_scan(driver, timeout=8)
        print_info("Clicked attachment Send button using preview scan.")
        return
    except TimeoutException as error:
        errors.append(short_error(error))

    try:
        click_send_button(driver, timeout=timeout)
        print_info("Clicked attachment Send button using Selenium selector.")
        return
    except TimeoutException as error:
        errors.append(short_error(error))

    try:
        click_attachment_send_by_screen_position(driver)
        print_info("Clicked attachment Send button using screen-position fallback.")
        return
    except Exception as error:
        errors.append(short_error(error))

    raise TimeoutException("Attachment Send click failed: " + " | ".join(errors))


def send_with_all_fallbacks(
    driver: webdriver.Chrome,
    active_element=None,
    timeout: int = 20,
    prefer_preview_button: bool = False,
) -> None:
    """Send using Selenium selectors, JS scan, then ENTER key fallbacks."""
    if prefer_preview_button:
        try:
            click_send_button_by_js_scan(driver, timeout=8)
            return
        except TimeoutException:
            pass

    if not prefer_preview_button:
        try:
            click_send_button(driver, timeout=timeout)
            return
        except TimeoutException:
            pass

        try:
            click_send_button_by_js_scan(driver, timeout=8)
            return
        except TimeoutException:
            pass
    else:
        try:
            click_send_button(driver, timeout=timeout)
            return
        except TimeoutException:
            pass

    if active_element is not None:
        try:
            active_element.send_keys(Keys.ENTER)
            return
        except WebDriverException:
            pass

    try:
        pyautogui.press("enter")
        return
    except Exception as error:
        raise TimeoutException(f"Could not click Send or press ENTER: {short_error(error)}")


def wait_for_attachment_send_button(driver: webdriver.Chrome, timeout: int = 30) -> None:
    """Wait until the attachment preview has a Send button available."""
    end_time = time.time() + timeout
    while time.time() < end_time:
        for by, selector in send_button_selectors():
            try:
                if any(element.is_displayed() and element.is_enabled() for element in driver.find_elements(by, selector)):
                    return
            except WebDriverException:
                continue
        try:
            js_found = driver.execute_script(
                """
                return Array.from(document.querySelectorAll(
                    'button[aria-label="Send"],[role="button"][aria-label="Send"],[data-testid="send"],[data-icon*="send"],[aria-label*="Send"],[title*="Send"]'
                )).some((node) => {
                    const clickable = node.closest('button,[role="button"]') || node;
                    const rect = clickable.getBoundingClientRect();
                    const style = window.getComputedStyle(clickable);
                    return rect.width > 0 &&
                        rect.height > 0 &&
                        style.visibility !== 'hidden' &&
                        style.display !== 'none' &&
                        clickable.getAttribute('aria-disabled') !== 'true';
                });
                """
            )
            if js_found:
                return
        except WebDriverException:
            pass
        time.sleep(0.7)

    raise TimeoutException("Attachment Send button did not appear.")


def wait_for_attachment_preview(driver: webdriver.Chrome, attachment: Path, timeout: int = 45) -> None:
    """Wait until WhatsApp's attachment preview/upload screen is actually ready."""
    file_name = attachment.name.lower()
    end_time = time.time() + timeout

    preview_script = """
        const fileName = arguments[0];
        const bodyText = (document.body.innerText || '').toLowerCase();
        const selectors = [
            '[role="dialog"]',
            'div[aria-label*="Preview"]',
            'div[aria-label*="preview"]',
            'img[src^="blob:"]',
            'video[src^="blob:"]',
            '[data-icon="media-editor"]',
            '[data-icon="document"]',
            '[data-icon="x-viewer"]'
        ];

        const hasPreviewNode = selectors.some((selector) =>
            document.querySelector(selector)
        );

        const hasFileName = fileName && bodyText.includes(fileName);
        const hasUploadingText =
            bodyText.includes('uploading') ||
            bodyText.includes('preparing') ||
            bodyText.includes('send media') ||
            bodyText.includes('add a caption') ||
            bodyText.includes('caption');

        return Boolean(hasPreviewNode || hasFileName || hasUploadingText);
    """

    while time.time() < end_time:
        try:
            if driver.execute_script(preview_script, file_name):
                return
        except WebDriverException:
            pass
        time.sleep(0.7)

    raise TimeoutException("Attachment preview did not open after selecting file.")


def send_text_message(driver: webdriver.Chrome, number: str, message: str) -> None:
    """Send message by real clipboard copy/paste only. No template typing."""
    chat_box = wait_for_chat_box(driver, number)
    chat_box.click()
    time.sleep(random.uniform(0.5, 1.3))

    paste_text_into_box(driver, chat_box, message)

    time.sleep(random.uniform(0.4, 1.1))
    send_with_all_fallbacks(driver, active_element=chat_box, timeout=12)

    time.sleep(random.uniform(1.5, 2.5))


def choose_file_input(driver: webdriver.Chrome, attachment: Path):
    """Choose the correct hidden file input for images or documents."""
    extension = attachment.suffix.lower()
    file_inputs = driver.find_elements(By.XPATH, "//input[@type='file']")
    if not file_inputs:
        raise NoSuchElementException("File input not found after clicking attach.")

    print_info(f"Found {len(file_inputs)} WhatsApp file input(s).")

    if extension in {".jpg", ".jpeg", ".png"}:
        for file_input in file_inputs:
            accept = (file_input.get_attribute("accept") or "").lower()
            if "image" in accept or "video" in accept:
                print_info(f"Using media upload input: accept='{accept or '*'}'")
                return file_input

    if extension == ".pdf":
        for file_input in file_inputs:
            accept = (file_input.get_attribute("accept") or "").lower()
            if "image" not in accept and "video" not in accept and "audio" not in accept:
                print_info(f"Using document upload input: accept='{accept or '*'}'")
                return file_input

    print_info("Using last available upload input as fallback.")
    return file_inputs[-1]


def send_attachment(driver: webdriver.Chrome, caption: str) -> bool:
    """Send the configured jpg, png, or pdf. Return True if text was sent as caption."""
    attachment = resolve_attachment_path(ATTACHMENT_PATH).resolve()

    attach_button_selectors = [
        (By.XPATH, "//button[@title='Attach']"),
        (By.XPATH, "//button[@aria-label='Attach']"),
        (By.XPATH, "//div[@title='Attach']"),
        (By.XPATH, "//div[@aria-label='Attach']"),
        (By.XPATH, "//span[@data-icon='plus']/ancestor::button"),
        (By.XPATH, "//span[@data-icon='plus']/ancestor::*[@role='button']"),
        (By.XPATH, "//span[@data-icon='clip']/ancestor::button"),
        (By.XPATH, "//span[@data-icon='clip']/ancestor::*[@role='button']"),
    ]
    click_first_available(driver, attach_button_selectors, timeout=12)
    time.sleep(random.uniform(0.8, 1.4))

    print_info(f"Uploading attachment: {attachment}")
    upload_input = choose_file_input(driver, attachment)
    upload_input.send_keys(str(attachment))
    wait_for_attachment_preview(driver, attachment, timeout=45)
    wait_for_attachment_send_button(driver, timeout=35)
    time.sleep(random.uniform(0.8, 1.5))

    caption_selectors = [
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true' and @role='textbox']"),
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true' and contains(@aria-label, 'caption')]"),
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true' and contains(@aria-placeholder, 'caption')]"),
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true' and contains(@aria-label, 'Caption')]"),
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true' and contains(@aria-placeholder, 'Caption')]"),
        (By.XPATH, "//*[@contenteditable='true' and contains(@aria-label, 'caption')]"),
        (By.XPATH, "//*[@contenteditable='true' and contains(@aria-placeholder, 'caption')]"),
        (By.XPATH, "//*[@contenteditable='true' and contains(@aria-label, 'Caption')]"),
        (By.XPATH, "//*[@contenteditable='true' and contains(@aria-placeholder, 'Caption')]"),
        (By.XPATH, "//div[@role='dialog']//footer//div[@contenteditable='true']"),
        (By.XPATH, "//div[@role='dialog']//div[@contenteditable='true']"),
    ]

    caption_added = False
    if SEND_TEXT_AS_ATTACHMENT_CAPTION:
        try:
            caption_box = find_first_visible(driver, caption_selectors, timeout=12)
            paste_text_into_box(driver, caption_box, caption)
            caption_added = True
        except TimeoutException:
            print_info("No caption field found for this attachment. Sending file first, then text separately.")
    else:
        print_info("Sending attachment first. Text will be sent separately after file upload.")

    time.sleep(random.uniform(0.5, 1.2))
    click_attachment_preview_send(driver, timeout=25)
    time.sleep(random.uniform(4.0, 6.0))
    return caption_added


def send_to_number(driver: webdriver.Chrome, number: str, template: str) -> tuple[str, str]:
    """Send text and optional attachment. Return status and reason."""
    try:
        if ENABLE_ATTACHMENT:
            wait_for_chat_box(driver, number)
            caption_added = send_attachment(driver, template)
            if not caption_added:
                send_text_message(driver, number, template)
        else:
            send_text_message(driver, number, template)

        return "Sent", "-"

    except ValueError as error:
        return "Failed", str(error)
    except TimeoutException as error:
        return "Failed", f"Timeout: {error}"
    except WebDriverException as error:
        return "Failed", f"Browser error: {short_error(error)}"
    except Exception as error:
        return "Failed", short_error(error)


def choose_template(previous_template_number: int | None) -> tuple[int, str]:
    """Pick a random template while avoiding an immediate repeat when possible."""
    available_numbers = list(range(1, len(TEMPLATES) + 1))
    if previous_template_number in available_numbers and len(available_numbers) > 1:
        available_numbers.remove(previous_template_number)

    template_number = random.choice(available_numbers)
    return template_number, TEMPLATES[template_number - 1]


def short_error(error: Exception) -> str:
    """Keep console/report errors readable."""
    text = str(error).replace("\n", " ").strip()
    return text[:180] if text else error.__class__.__name__


def build_report_rows(results: list[dict]) -> list[dict]:
    """Convert raw result rows into the requested emoji report format."""
    status_map = {
        "Sent": "✅ Sent",
        "Failed": "❌ Failed",
        "Skipped": "⚠ Skipped",
    }

    rows = []
    for result in results:
        rows.append(
            {
                "Number": result["number"],
                "Status": status_map.get(result["status"], result["status"]),
                "Time": result["time"],
                "Template": result["template"],
                "Reason": result["reason"],
            }
        )
    return rows


def generate_excel_report(results: list[dict]) -> None:
    """Create report.xlsx with readable columns and summary totals."""
    report_path = BASE_DIR / REPORT_FILE
    rows = build_report_rows(results)
    df = pd.DataFrame(rows, columns=["Number", "Status", "Time", "Template", "Reason"])

    sent_count = sum(1 for row in results if row["status"] == "Sent")
    failed_count = sum(1 for row in results if row["status"] == "Failed")
    skipped_count = sum(1 for row in results if row["status"] == "Skipped")

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", index=False)

    workbook = load_workbook(report_path)
    sheet = workbook["Report"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    summary_fill = PatternFill("solid", fgColor="D9EAD3")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    summary_start = sheet.max_row + 3
    summary_rows = [
        ("Total Sent", sent_count),
        ("Total Failed", failed_count),
        ("Total Skipped", skipped_count),
    ]

    sheet.cell(row=summary_start, column=1, value="Summary")
    sheet.cell(row=summary_start, column=1).font = Font(bold=True)

    for offset, (label, value) in enumerate(summary_rows, start=1):
        label_cell = sheet.cell(row=summary_start + offset, column=1, value=label)
        value_cell = sheet.cell(row=summary_start + offset, column=2, value=value)
        label_cell.fill = summary_fill
        value_cell.fill = summary_fill
        label_cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 55)

    workbook.save(report_path)
    print_info(f"Excel report saved: {report_path}")


def sleep_between_messages() -> None:
    delay = random.randint(int(MIN_DELAY), int(MAX_DELAY))
    print_info(f"Waiting {delay} seconds")
    time.sleep(delay)


def main() -> None:
    collect_runtime_settings()
    validate_config()

    contacts = read_contacts()
    print_info(f"Loaded {len(contacts)} contacts")

    sent_before = load_sent_numbers_from_logs()
    if sent_before:
        print_info(f"Resume enabled. Found {len(sent_before)} previously sent numbers.")

    sent_before = show_queue_details(contacts, sent_before)

    driver = create_driver()
    results: list[dict] = []
    sent_today = 0
    previous_template_number: int | None = None

    try:
        wait_for_login(driver)
        input("[INFO] Press ENTER to start sending messages...")

        for index, number in enumerate(contacts):
            if sent_today >= DAILY_LIMIT:
                print_info(f"Daily limit reached: {DAILY_LIMIT}")
                break

            result_time = now_text()
            template_number, template = choose_template(previous_template_number)
            previous_template_number = template_number
            if template_number <= len(TEMPLATE_NAMES):
                template_label = f"Template {template_number}: {TEMPLATE_NAMES[template_number - 1]}"
            else:
                template_label = f"Template {template_number}"

            if number in sent_before:
                print_failed(f"{number} skipped: Already Sent")
                append_log(number, "Skipped", template_label, "Already Sent")
                results.append(
                    {
                        "number": number,
                        "status": "Skipped",
                        "time": result_time,
                        "template": template_label,
                        "reason": "Already Sent",
                    }
                )
                continue

            if not is_basic_number_format_valid(number):
                print_failed(f"{number} skipped: Invalid local number format")
                append_log(number, "Failed", template_label, "Invalid local number format")
                results.append(
                    {
                        "number": number,
                        "status": "Failed",
                        "time": result_time,
                        "template": template_label,
                        "reason": "Invalid local number format",
                    }
                )
                continue

            print_info(f"Sending to {number}")
            print_info(f"Template Selected: #{template_number}")

            status, reason = send_to_number(driver, number, template)
            result_time = now_text()
            append_log(number, status, template_label, reason)

            results.append(
                {
                    "number": number,
                    "status": status,
                    "time": result_time,
                    "template": template_label,
                    "reason": reason,
                }
            )

            if status == "Sent":
                sent_today += 1
                sent_before.add(number)
                print_success("Message sent")
            else:
                print_failed(reason)

            if sent_today < DAILY_LIMIT and index < len(contacts) - 1:
                sleep_between_messages()

    except KeyboardInterrupt:
        print_info("CTRL+C detected. Saving report before exit...")
    finally:
        if results:
            generate_excel_report(results)
        else:
            print_info("No results to report.")

        try:
            driver.quit()
        except Exception:
            pass

        print_info("Bot stopped.")


if __name__ == "__main__":
    try:
        main()
    except Exception as fatal_error:
        print_failed(short_error(fatal_error))
        sys.exit(1)
